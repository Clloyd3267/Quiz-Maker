###################################################################################################
# Name        : QuestionList.py
# Author(s)   : Chris Lloyd, Andrew Southwick
# Description : Class to store and manage questions
###################################################################################################

# External Imports
from pathlib import Path # Used for file manipulation
import openpyxl # For reading in questions

# Project Imports
from MaterialList import * # Used to check the reference of questions


class QuestionList:
    """
    A class to store the questions and perform functions on them.

    Attributes:
        mL (MaterialList): Used for checking the references.
        importErrors (array of errors): Errors while importing.
        questionDatabase (array of Question): The array that stores all of the questions.
        isGospel (bool): Variable to store whether year is a gospel or not.
    """

    mL = MaterialList() # Object of type MaterialList
    importErrors = [] # Errors while importing
    questionDatabase = []  # The array that stores all of the questions
    isGospel = 0  # Variable to store whether year is a gospel or not

    def __init__(self, questionFileName = "Questions.xlsx"):
        """
        The constructor for class Question List.

        Parameters:
            questionFileName (str): The input filename for questions (Defaults to "Questions.xlsx").
        """

        self.importQuestions(questionFileName)
        self.exportQuestions(questionFileName)

    ####################################################################################################################
    # Helper Functions
    ####################################################################################################################
    def importQuestions(self, questionFileName):
        """
        Function to import questions and populate QuestionList object.

        Parameters:
            questionFileName (str): The input filename for questions.
        """

        dataFilePath = Path("../Data Files/")  # Path where datafiles are stored
        if questionFileName == "Questions.xlsx":
            questionFilePath = dataFilePath / questionFileName
        else:
            questionFilePath = questionFileName

        try:
            book = openpyxl.load_workbook(questionFilePath)
        except IOError:
            print("Error => Question file does not exist!!!")
            return

        sheet = book.worksheets[0]  # Open the first sheet
        for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 7):
            question = []
            for cell in row:
                if not cell.value:
                    question.append("")
                else:
                    question.append(str(cell.value))

            tempQuestionBook = ""
            for char in question[0]:
                if char.isalnum():
                    tempQuestionBook += char
            question[0] = tempQuestionBook

            question[1] = question[1].replace(" ", "")
            question[2] = question[2].replace(" ", "")
            question[3] = question[3].replace(" ", "")


            # Check reference
            if not question[0]:
                self.importErrors.append(["Error => No book!!!", question[:7]])
                continue
            if not question[1]:
                self.importErrors.append(["Error => No chapter!!!", question[:7]])
                continue
            if not question[2]:
                self.importErrors.append(["Error => No verse start!!!", question[:7]])
                continue
            if not self.mL.checkRef([question[0], question[1], question[2]]):
                self.importErrors.append(["Error => Verse start invalid!!!", question[:7]])
                continue
            if question[3] != ""  and not self.mL.checkRef([question[0], question[1], question[3]]):
                self.importErrors.append(["Error => Verse end invalid!!!", question[:7]])
                continue

            # Check type
            searchTypes = \
                {"INT": ["INT", "INTF"],
                 "CR": ["CR", "CRMA"],
                 "CVR": ["CVR", "CVRMA"],
                 "MA": ["MA"],
                 "Q": ["Q", "Q2"],
                 "FTV": ["FTV", "FT2V", "F2V", "FT", "FTN"],
                 "SIT": ["SIT"]}
            valid = False
            for qMainType in searchTypes.keys():
                for qType in searchTypes[qMainType]:
                    if question[4].lower().find(qType.lower()) != -1:
                        # Check for Gospel
                        if qMainType == "SIT":
                            self.isGospel = 1
                        valid = True
            if not valid:
                self.importErrors.append(["Error => Question type invalid!!!", question[:7]])
                continue

            # Check question and answer
            if not question[5]:
                self.importErrors.append(["Error => No question!!!", question[:7]])
                continue
            if not question[6]:
                self.importErrors.append(["Error => No answer!!!", question[:7]])
                continue

            # questionObj = Question(question[0], question[1], question[2],
                                   # question[3], question[4], question[5], question[6])
            self.questionDatabase.append(question)
        book.close()

    def exportQuestions(self, questionFileName):
        """
        Function to export questions from QuestionList object.

        Parameters:
            questionFileName (str): The output filename for questions.
        """

        dataFilePath = Path("../Data Files/")  # Path where datafiles are stored
        if questionFileName == "Questions.xlsx":
            questionFilePath = dataFilePath / questionFileName
        else:
            questionFilePath = questionFileName

        try:
            book = openpyxl.load_workbook(questionFilePath)
        except IOError:
            print("Error => Question file does not exist!!!")
            return
        sheet1 = book.active  # Open the active sheet questionFileName
        book.remove(sheet1)
        for i in book.worksheets:
            if i.title == "Errors":
                book.remove(i)
            elif i.title == "Questions":
                book.remove(i)
        book.create_sheet("Questions")
        book.create_sheet("Errors")
        sheet1 = book["Questions"]
        sheet2 = book["Errors"]

        # Add headers
        sheet1.cell(row = 1, column = 1).value = "Book"
        sheet1.cell(row = 1, column = 2).value = "Chapter"
        sheet1.cell(row = 1, column = 3).value = "Verse Start"
        sheet1.cell(row = 1, column = 4).value = "Verse End"
        sheet1.cell(row = 1, column = 5).value = "Type"
        sheet1.cell(row = 1, column = 6).value = "Question"
        sheet1.cell(row = 1, column = 7).value = "Answer"

        for rowNum, question in enumerate(self.questionDatabase, start = 2):
            sheet1.cell(row = rowNum, column = 1).value = question[0]
            sheet1.cell(row = rowNum, column = 2).value = question[1]
            sheet1.cell(row = rowNum, column = 3).value = question[2]
            sheet1.cell(row = rowNum, column = 4).value = question[3]
            sheet1.cell(row = rowNum, column = 5).value = question[4]
            sheet1.cell(row = rowNum, column = 6).value = question[5]
            sheet1.cell(row = rowNum, column = 7).value = question[6]

        for rowNum, questionError in enumerate(self.importErrors, start = 1):
            i = 1
            while i != 8:
                sheet2.cell(row = rowNum, column = i).value = questionError[1][i - 1]
                i += 1
            sheet2.cell(row = rowNum, column = 8).value = questionError[0]

        book.save(dataFilePath / questionFileName)












        # 1. Find all questions
        # 2. Fill Minimums
        # 3. Fill remaining numbered questions
        # 4. Add question numbers
        # 5. Add a and b questions
        # 6. Export quizzes